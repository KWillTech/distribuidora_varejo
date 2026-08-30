"""Geração e exportação dos relatórios."""
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4,landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle
from models.auth import Permission
from models.report import ReportType
from services.rbac import require_permission
class ReportService:
    def __init__(self,repository,audit):self.repository=repository; self.audit=audit
    def load(self,session,filters):
        require_permission(session,Permission.REPORTS_VIEW)
        if filters.report_type==ReportType.CREDIT:require_permission(session,Permission.CREDIT_VIEW)
        if filters.report_type==ReportType.COMMANDS:require_permission(session,Permission.TABS_HISTORY)
        return self.repository.load(filters)
    def export_excel(self,session,data,path):
        require_permission(session,Permission.REPORTS_EXPORT)
        if data.title.startswith("Fiado"):require_permission(session,Permission.CREDIT_EXPORT)
        wb=Workbook(); ws=wb.active; ws.title="Relatório"; ws.append([data.title]); ws.append(data.columns)
        for row in data.rows:ws.append(row)
        ws.append([]); ws.append([data.total_label,data.total_value]); ws["A1"].font=Font(size=16,bold=True,color="B45F06")
        for cell in ws[2]:cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="1A1A1A")
        for column in ws.columns:ws.column_dimensions[column[0].column_letter].width=min(45,max(len(str(c.value or "")) for c in column)+2)
        wb.save(path); self._audit(session,"relatorio_excel_exportado",path); return Path(path)
    def export_pdf(self,session,data,path):
        require_permission(session,Permission.REPORTS_EXPORT)
        if data.title.startswith("Fiado"):require_permission(session,Permission.CREDIT_EXPORT)
        doc=SimpleDocTemplate(str(path),pagesize=landscape(A4),rightMargin=24,leftMargin=24,topMargin=28,bottomMargin=34); styles=getSampleStyleSheet(); story=[Paragraph(data.title,styles["Title"]),Spacer(1,10)]; table=Table([data.columns]+data.rows,repeatRows=1); table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1A1A1A")),("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#FFB000")),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),("GRID",(0,0),(-1,-1),.25,colors.grey),("VALIGN",(0,0),(-1,-1),"TOP"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#FFF7E6")])])) ; story.extend([table,Spacer(1,12),Paragraph(f"<b>{data.total_label}: {data.total_value}</b>",styles["Normal"])]); issued=datetime.now().strftime("%d/%m/%Y %H:%M")
        def footer(canvas,document):
            canvas.saveState(); canvas.setFont("Helvetica",7); canvas.setFillColor(colors.grey); canvas.drawString(24,16,f"Adega do Bruninho - Emitido em {issued}"); canvas.drawRightString(landscape(A4)[0]-24,16,f"Página {document.page}"); canvas.restoreState()
        doc.build(story,onFirstPage=footer,onLaterPages=footer); self._audit(session,"relatorio_pdf_exportado",path); return Path(path)
    def _audit(self,session,action,path):self.audit.record(user=session.user,action=action,module="relatorios",details={"arquivo":str(path)})
